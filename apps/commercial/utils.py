import openpyxl
import requests
from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from django.db import transaction
from apps.commercial.models.instrument_models import Instrument
from apps.commercial.models.client_models import Branch, Client


def client_obtain_and_create(ruc):
    """Query the SUNAT API, create a client and its branches."""
    url_full = f"https://api.apis.net.pe/v2/sunat/ruc/full?numero={ruc}"
    url_basic = f"https://api.apis.net.pe/v2/sunat/ruc?numero={ruc}"
    token = "apis-token-12335.66f3s7HjN6Nel8gcQOHo0le531JKyeMB"
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {token}',
    }
    try:
        # Request data from both the full and basic URLs
        response_full = requests.get(url_full, headers=headers)
        response_full.encoding = 'utf-8'
        data_full = response_full.json() if response_full.status_code == 200 else {}

        response_basic = requests.get(url_basic, headers=headers)
        response_basic.encoding = 'utf-8'
        data_basic = response_basic.json() if response_basic.status_code == 200 else {}

        # Imprime las respuestas para depuración
        print("=== FULL DATA RESPONSE ===")
        print(data_full)
        print("=== BASIC DATA RESPONSE ===")
        print(data_basic)

        # Combinar datos para el cliente: data_full tiene prioridad
        data = {**data_basic, **data_full}

        # Campos del cliente usando la data combinada
        client_data = {
            'business_name': data.get('razonSocial', 'No disponible'),
            'tax_id': data.get('numeroDocumento', 'No disponible'),
            'address': data.get('direccion', 'No disponible'),
            'department': data.get('departamento', 'No disponible'),
            'province': data.get('provincia', 'No disponible'),
            'district': data.get('distrito', 'No disponible'),
            'economic_activity': data.get('actividadEconomica', 'No disponible'),
            'status': data.get('estado', 'No disponible'),
            'condition': data.get('condicion', 'No disponible'),
            'is_retention_agent': data.get('EsAgenteRetencion', False),
            "workers_number": data.get("numeroTrabajadores", "Not available"),
            "billing_type": data.get("tipoFacturacion", "Computerized"),
            "accounting_type": data.get("tipoContabilidad", "Computerized"),
            "foreign_trade": data.get("comercioExterior", "No Activity"),
            "company_type": data.get("tipo", "Limited Liability Company"),
        }

        # Crear el cliente y las sucursales en una transacción
        with transaction.atomic():
            client = Client.objects.create(**client_data)

            # Para las sucursales usamos los datos de 'localesAnexos' de data_basic
            locales_anexos = data_basic.get('localesAnexos', [])
            if isinstance(locales_anexos, list):
                for branch_data in locales_anexos:
                    Branch.objects.create(
                        client=client,
                        name=branch_data.get('direccion', 'Sucursal sin nombre'),
                        address=branch_data.get('direccion', ''),
                        department=branch_data.get('departamento', ''),
                        province=branch_data.get('provincia', ''),
                        district=branch_data.get('distrito', ''),
                        is_headquarters=False  # Ajusta según tu lógica
                    )
            else:
                print("No 'localesAnexos' found or it's not a list in data_basic.")

        return client
    except Exception as e:
        print(f"Error while obtaining and creating client with RUC {ruc}: {e}")
        return None


def process_instruments_excel(file):
    """
    Procesa un archivo Excel con 27 columnas (de A a AA) y guarda/actualiza registros
    en la tabla MeasurementInstrument.
    """

    # Funciones auxiliares para parsear valores
    def parse_bool(value):
        """Devuelve True si el texto equivale a 'yes', 'si', 'true', '1', de lo contrario False."""
        if not value:
            return False
        return str(value).strip().lower() in ['yes', 'si', 'true', '1']

    def parse_decimal(value):
        """Convierte a Decimal; si falla, devuelve Decimal('0.00')."""
        try:
            # Si vienen decimales con coma, como "123,45", los convertimos a punto "123.45"
            return Decimal(str(value).replace(',', '.'))
        except:
            return Decimal('0.00')

    def parse_int(value):
        """Convierte a entero; si falla, devuelve 1 por defecto."""
        try:
            return int(value)
        except:
            return 1

    try:
        # Carga el workbook y selecciona la hoja activa
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        # Iteramos desde la fila 2 (asumiendo la fila 1 son encabezados),
        # y de la columna 1 (A) hasta la 27 (AA).
        for row_num, row in enumerate(
                sheet.iter_rows(min_row=2, min_col=1, max_col=27, values_only=True),
                start=2
        ):
            # Si la fila está completamente vacía, la saltamos
            if not any(row):
                continue

            # Mapeamos cada columna con su campo en el modelo:
            # row[0]  -> Item (no lo usamos directamente, pero está en la hoja)
            discipline = str(row[1] or "").strip()  # B: Disciplina / Magnitud
            subdiscipline = str(row[2] or "").strip()  # C: Subdisiciplina
            instrument = str(row[3] or f"Desconocido {row_num}").strip()  # D: Instrumento
            procedure_code = str(row[4] or "").strip()  # E: Código Procedimiento
            procedure = str(row[5] or "").strip()  # F: Procedimiento
            commercial_area_notes = str(row[6] or "").strip()  # G: Observaciones para comercial
            measurement_range_from = str(row[7] or "").strip()  # H: Intervalo desde
            measurement_range_to = str(row[8] or "").strip()  # I: Intervalo hasta
            indicator_resolution = str(row[9] or "").strip()  # J: Resolución
            brand = str(row[10] or "").strip()  # K: Marca
            model = str(row[11] or "").strip()  # L: Modelo
            expanded_uncertainty = str(row[12] or "").strip()  # M: Incertidumbre
            accredited = parse_bool(row[13])  # N: Acreditado (bool)
            accreditation_condition = str(row[14] or "").strip()  # O: Condición de Acreditado
            employed_standards = str(row[15] or "").strip()  # P: Patrones empleados
            man_hours = str(row[16] or "").strip()  # Q: Horas hombre
            necessary_workers = parse_int(row[17])  # R: Personal necesario (int)
            calibration_cost = parse_decimal(row[18])  # S: Costo calibración
            verification_cost = parse_decimal(row[19])  # T: Costo verificación
            preventive_maintenance_cost = parse_decimal(row[20])  # U: Costo mant. preventivo
            transport_cost = parse_decimal(row[21])  # V: Costo transporte
            for_sale = parse_bool(row[22])  # W: VENTA (bool)
            technical_notes = str(row[23] or "").strip()  # X: Observaciones técnicas
            cost_center = str(row[24] or "").strip()  # Y: Centro de costos
            in_arequipa = parse_bool(row[25])  # Z: AREQUIPA (bool)
            in_lima = parse_bool(row[26])  # AA: LIMA (bool)

            # Creamos o actualizamos el registro según 'instrument' como identificador único
            Instrument.objects.update_or_create(
                name=instrument,
                measurement_range_from=measurement_range_from,
                measurement_range_to=measurement_range_to,
                defaults={
                    'discipline': discipline,
                    'subdiscipline': subdiscipline,
                    'procedure_code': procedure_code,
                    'procedure': procedure,
                    'commercial_area_notes': commercial_area_notes,
                    'indicator_resolution': indicator_resolution,
                    'expanded_uncertainty': expanded_uncertainty,
                    'accredited': accredited,
                    'accreditation_condition': accreditation_condition,
                    'employed_standards': employed_standards,
                    'man_hours': man_hours,
                    'necessary_workers': necessary_workers,
                    'calibration_cost': calibration_cost,
                    'verification_cost': verification_cost,
                    'preventive_maintenance_cost': preventive_maintenance_cost,
                    'transport_cost': transport_cost,
                    'for_sale': for_sale,
                    'technical_notes': technical_notes,
                    'cost_center': cost_center,
                    'in_arequipa': in_arequipa,
                    'in_lima': in_lima,
                }
            )

        return "Archivo procesado correctamente."

    except Exception as e:
        return f"Error al procesar el archivo: {str(e)}"


def export_instruments_to_excel():
    # Crea un objeto Workbook de openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Instrumentos"

    # Define los encabezados de la tabla
    headers = [
        'Disciplina', 'Subdisciplina', 'Instrumento', 'Código Procedimiento',
        'Procedimiento', 'Observaciones para comercial', 'Intervalo desde', 'Intervalo hasta',
        'Resolución', 'Marca', 'Modelo', 'Incertidumbre', 'Acreditado', 'Condición de Acreditado',
        'Patrones empleados', 'Horas hombre', 'Personal necesario', 'Costo calibración',
        'Costo verificación', 'Costo mant. preventivo', 'Costo transporte', 'Venta', 'Observaciones técnicas',
        'Centro de costos', 'Arequipa', 'Lima'
    ]

    # Escribe los encabezados en la primera fila
    ws.append(headers)

    # Obtén los datos de la base de datos (simulación de consulta a la BD)
    instruments = Instrument.objects.all()

    # Agrega los datos de los instrumentos en las filas siguientes
    for instrument in instruments:
        row = [
            instrument.discipline, instrument.subdiscipline, instrument.name,
            instrument.procedure_code, instrument.procedure, instrument.commercial_area_notes,
            instrument.measurement_range_from, instrument.measurement_range_to,
            instrument.indicator_resolution, instrument.brand, instrument.model,
            instrument.expanded_uncertainty, instrument.accredited, instrument.accreditation_condition,
            instrument.employed_standards, instrument.man_hours, instrument.necessary_workers,
            instrument.calibration_cost, instrument.verification_cost,
            instrument.preventive_maintenance_cost, instrument.transport_cost,
            instrument.for_sale, instrument.technical_notes, instrument.cost_center,
            instrument.in_arequipa, instrument.in_lima
        ]
        ws.append(row)

    # Guardar el archivo en memoria usando BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Necesario para volver al principio del archivo en memoria

    # Cerrar el archivo para evitar problemas
    wb.close()

    return excel_file
